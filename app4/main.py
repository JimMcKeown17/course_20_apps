import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path
from datetime import datetime

filepaths = glob.glob("invoices/*.xlsx")
for filepath in filepaths:

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    filename = Path(filepath).stem
    invoice_nr, date = filename.split("-")
    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Invoice nr. {invoice_nr}", ln=1)

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"{date}", ln=1)

    #Add Headers
    columns = df.columns
    columns = [item.replace("_", " ").title() for item in columns]
    pdf.set_font(family="Times", size=12, style="B")
    pdf.cell(w=30, h=8, txt=str(columns[0]), border=1)
    pdf.cell(w=60, h=8, txt=str(columns[1]), border=1)
    pdf.cell(w=40, h=8, txt=str(columns[2]), border=1)
    pdf.cell(w=30, h=8, txt=str(columns[3]), border=1)
    pdf.cell(w=30, h=8, txt=str(columns[4]), border=1, ln=1)

    total = 0
    for index, row in df.iterrows():

        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80,80,80)
        pdf.cell(w=30,h=8,txt=str(row["product_id"]), border=1)
        pdf.cell(w=60,h=8,txt=str(row["product_name"]), border=1)
        pdf.cell(w=40,h=8,txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30,h=8,txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30,h=8,txt=str(row["total_price"]), border=1, ln=1)
        total += row["total_price"]
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=60, h=8, txt="", border=1)
    pdf.cell(w=40, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=str(total), border=1, ln=1)
    pdf.cell(w=30, h=20, ln=2)

    pdf.set_font(family="Times", style="B")
    pdf.cell(w=100, h=8, txt=f"The total due amount is {total} Euros.")


    pdf.output(f"PDFs/{filename}.pdf")